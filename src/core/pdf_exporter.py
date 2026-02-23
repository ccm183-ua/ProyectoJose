"""
Exportador de presupuestos Excel a PDF.

Usa win32com.client (Microsoft Excel COM automation) para generar un PDF
idéntico al "Imprimir como PDF" desde Excel. Requiere Microsoft Excel instalado.

Funcionalidades de paginación:
- Inserta físicamente las filas del encabezado al inicio de cada página
  intermedia (2..n-1) usando COM en un subprocess aislado, para que
  aparezcan tanto en el xlsx como en el PDF sin recurrir a PrintTitleRows.
- Fuerza un salto de página antes de la sección de resumen para que siempre
  aparezca al inicio de la última hoja.
- La última página (resumen) queda limpia, sin encabezado repetido.
"""

import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from typing import Optional, Tuple

logger = logging.getLogger(__name__)

_SHEET_XML = "xl/worksheets/sheet1.xml"
_WORKBOOK_XML = "xl/workbook.xml"


class PDFExporter:
    """Exporta un .xlsx a PDF usando Excel COM (Windows)."""

    @staticmethod
    def is_available() -> bool:
        """True si Microsoft Excel está disponible en el sistema."""
        try:
            import win32com.client  # noqa: F401
            return True
        except ImportError:
            return False

    # ------------------------------------------------------------------
    # Detección de filas con openpyxl (read-only, no modifica el archivo)
    # ------------------------------------------------------------------

    @staticmethod
    def _find_obra_rows(xlsx_path: str) -> Tuple[Optional[int], Optional[int]]:
        """Escanea columna A buscando celdas que contengan "Obra:".

        Returns:
            (header_row, summary_row) donde header_row es la primera "Obra:"
            en filas 1-20 y summary_row es la última "Obra:" en toda la hoja.
            Cualquiera puede ser None si no se encuentra.
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active

            header_row: Optional[int] = None
            last_obra: Optional[int] = None

            for row in ws.iter_rows(min_col=1, max_col=1):
                cell = row[0]
                if cell.value and "Obra:" in str(cell.value):
                    if header_row is None and cell.row <= 20:
                        header_row = cell.row
                    last_obra = cell.row

            wb.close()

            summary_row = last_obra if last_obra != header_row else None
            logger.info(
                "PDF export - filas detectadas: header=%s, summary=%s (last_obra=%s)",
                header_row, summary_row, last_obra,
            )
            return (header_row, summary_row)
        except Exception as exc:
            logger.warning("PDF export - error al leer xlsx con openpyxl: %s", exc)
            return (None, None)

    # ------------------------------------------------------------------
    # Inyección XML de configuración de página
    # ------------------------------------------------------------------

    @staticmethod
    def _inject_row_breaks(sheet_xml: str, summary_row: int) -> str:
        """Inserta un <rowBreaks> manual antes de *summary_row* en el XML.

        El atributo ``id`` del ``<brk>`` es 0-based en OOXML: indica la fila
        *después* de la cual se coloca el salto.  Para que *summary_row*
        (1-based) inicie en página nueva, usamos ``id = summary_row - 1``.
        """
        sheet_xml = re.sub(
            r"<rowBreaks[^>]*>.*?</rowBreaks>", "", sheet_xml, flags=re.DOTALL,
        )
        brk_id = summary_row - 1
        brk = (
            f'<rowBreaks count="1" manualBreakCount="1">'
            f'<brk id="{brk_id}" max="16383" man="1"/>'
            f'</rowBreaks>'
        )
        if "</headerFooter>" in sheet_xml:
            return sheet_xml.replace("</headerFooter>", f"</headerFooter>{brk}")
        if "<headerFooter/>" in sheet_xml:
            return sheet_xml.replace("<headerFooter/>", f"<headerFooter/>{brk}")
        if "<drawing " in sheet_xml:
            return sheet_xml.replace("<drawing ", f"{brk}<drawing ")
        return sheet_xml.replace("</worksheet>", f"{brk}</worksheet>")

    @staticmethod
    def _inject_print_titles(wb_xml: str, header_row: int) -> str:
        """Inserta un <definedNames> con _xlnm.Print_Titles en workbook.xml."""
        wb_xml = re.sub(
            r"<definedNames>.*?</definedNames>", "", wb_xml, flags=re.DOTALL,
        )
        m = re.search(r'<sheet name="([^"]+)"', wb_xml)
        sheet_name = m.group(1) if m else "Sheet1"
        end_row = header_row + 2
        dn = (
            f"<definedNames>"
            f'<definedName name="_xlnm.Print_Titles" localSheetId="0">'
            f"'{sheet_name}'!${header_row}:${end_row}"
            f"</definedName>"
            f"</definedNames>"
        )
        if "</sheets>" in wb_xml:
            return wb_xml.replace("</sheets>", f"</sheets>{dn}")
        return wb_xml.replace("</workbook>", f"{dn}</workbook>")

    @classmethod
    def apply_page_config(
        cls,
        xlsx_path: str,
        header_row: Optional[int] = None,
        summary_row: Optional[int] = None,
    ) -> bool:
        """Escribe el salto de página (rowBreaks) directamente en el xlsx.

        Solo inyecta el salto de página antes del resumen. Los PrintTitleRows
        NO se escriben en el xlsx porque Excel los aplicaría a TODAS las
        páginas (incluida la del resumen). Los encabezados repetidos se
        gestionan exclusivamente vía COM durante la exportación a PDF, donde
        sí podemos excluir la última página.

        Args:
            xlsx_path: Ruta al xlsx a modificar.
            header_row: Se ignora (reservado para la exportación PDF vía COM).
            summary_row: Fila de inicio del resumen para insertar page break.

        Returns:
            True si se modificó el archivo, False si no hacía falta o hubo error.
        """
        if summary_row is None:
            return False

        try:
            with zipfile.ZipFile(xlsx_path, "r") as z_in:
                namelist = z_in.namelist()
                contents = {n: z_in.read(n) for n in namelist}

            if _SHEET_XML not in contents:
                return False

            sheet_xml = contents[_SHEET_XML].decode("utf-8")
            sheet_xml = cls._inject_row_breaks(sheet_xml, summary_row)
            contents[_SHEET_XML] = sheet_xml.encode("utf-8")

            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            try:
                with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z_out:
                    for name in namelist:
                        z_out.writestr(name, contents[name])
                shutil.move(tmp_path, xlsx_path)
            except Exception:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                raise

            logger.info(
                "Page config aplicada al xlsx: %s (rowBreak antes de fila %s)",
                xlsx_path, summary_row,
            )
            return True

        except Exception as exc:
            logger.warning("Error al aplicar page config al xlsx: %s", exc)
            return False

    # ------------------------------------------------------------------
    # Inserción física de encabezados en saltos de página (subprocess)
    # ------------------------------------------------------------------

    # Script Python que se ejecuta en un proceso aislado para evitar
    # conflictos COM con el hilo principal de Qt.
    _COM_WORKER_SCRIPT = r'''
import json, os, sys

def main():
    xlsx = sys.argv[1]
    header_row = int(sys.argv[2])
    summary_row = int(sys.argv[3]) if sys.argv[3] != "0" else None
    NUM_HDR = 3

    import pythoncom, win32com.client
    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(xlsx)
        ws = wb.Worksheets(1)

        breaks = []
        for i in range(ws.HPageBreaks.Count):
            row = ws.HPageBreaks(i + 1).Location.Row
            if summary_row and row >= summary_row:
                continue
            breaks.append(row)

        if not breaks:
            wb.Close(SaveChanges=False)
            print(json.dumps({"ok": True, "inserted": 0, "breaks": []}))
            return

        breaks.sort(reverse=True)
        source = ws.Range(f"{header_row}:{header_row + NUM_HDR - 1}")

        for brk_row in breaks:
            source.Copy()
            ws.Range(f"{brk_row}:{brk_row + NUM_HDR - 1}").Insert(Shift=-4121)

        ws.ResetAllPageBreaks()
        breaks_asc = sorted(breaks)
        manual = []
        for idx, orig in enumerate(breaks_asc):
            pos = orig + idx * NUM_HDR
            ws.HPageBreaks.Add(ws.Cells(pos, 1))
            manual.append(pos)

        if summary_row:
            ns = summary_row + len(breaks) * NUM_HDR
            ws.HPageBreaks.Add(ws.Cells(ns, 1))
            manual.append(ns)

        wb.Save()
        wb.Close()
        print(json.dumps({"ok": True, "inserted": len(breaks), "breaks": manual}))
    except Exception as exc:
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        print(json.dumps({"ok": False, "error": str(exc)}))
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()

main()
'''

    @classmethod
    def insert_headers_at_page_breaks(cls, xlsx_path: str) -> bool:
        """Copia las filas del encabezado al inicio de cada página intermedia.

        Ejecuta las operaciones COM en un **subprocess** aislado para evitar
        que ``excel.Quit()`` / ``CoUninitialize()`` corrompan el estado COM
        del hilo principal de Qt (PySide6).

        Copia el archivo a una ruta temporal corta para evitar el límite
        MAX_PATH de 260 caracteres de Windows (las rutas de presupuestos
        pueden superar fácilmente ese límite).
        """
        header_row, summary_row = cls._find_obra_rows(xlsx_path)
        if not header_row:
            logger.info("insert_headers: sin header_row, nada que hacer")
            return False

        xlsx_abs = os.path.abspath(xlsx_path)
        summary_arg = str(summary_row) if summary_row else "0"

        fd, tmp_xlsx = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            shutil.copy2(xlsx_abs, tmp_xlsx)

            result = subprocess.run(
                [
                    sys.executable, "-c", cls._COM_WORKER_SCRIPT,
                    tmp_xlsx, str(header_row), summary_arg,
                ],
                capture_output=True,
                text=True,
                timeout=120,
            )

            if result.returncode != 0:
                logger.error(
                    "insert_headers: subprocess falló (rc=%d): %s",
                    result.returncode, result.stderr.strip(),
                )
                return False

            data = json.loads(result.stdout.strip())

            if data.get("ok") and data["inserted"] > 0:
                shutil.move(tmp_xlsx, xlsx_abs)
                logger.info(
                    "insert_headers: %d encabezados insertados, breaks=%s (%s)",
                    data["inserted"], data["breaks"], xlsx_path,
                )
                return True

            if not data.get("ok"):
                logger.error(
                    "insert_headers: worker error - %s", data.get("error"),
                )
            return False

        except subprocess.TimeoutExpired:
            logger.error("insert_headers: timeout (120s) en %s", xlsx_path)
            return False
        except Exception as exc:
            logger.error("insert_headers: error - %s [%s]", exc, type(exc).__name__)
            return False
        finally:
            if os.path.exists(tmp_xlsx):
                try:
                    os.unlink(tmp_xlsx)
                except OSError:
                    pass

    # ------------------------------------------------------------------
    # Exportación
    # ------------------------------------------------------------------

    def export(self, xlsx_path: str, pdf_path: str = None) -> Tuple[bool, str]:
        """Exporta un xlsx a PDF.

        Asume que ``insert_headers_at_page_breaks`` ya insertó los
        encabezados físicamente y que ``apply_page_config`` configuró el
        salto antes del resumen.  La exportación es un único paso directo.
        """
        if not os.path.exists(xlsx_path):
            return (False, f"El archivo no existe: {xlsx_path}")

        xlsx_path = os.path.abspath(xlsx_path)
        if not pdf_path:
            base, _ = os.path.splitext(xlsx_path)
            pdf_path = base + ".pdf"
        pdf_path = os.path.abspath(pdf_path)

        if not self.is_available():
            return (False, "win32com no está disponible. Instala pywin32.")

        logger.info("PDF export - inicio: %s -> %s", xlsx_path, pdf_path)

        com_initialized = False
        excel = None
        wb = None
        try:
            try:
                import pythoncom
                pythoncom.CoInitialize()
                com_initialized = True
            except Exception:
                pass

            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(xlsx_path)
            wb.ExportAsFixedFormat(0, pdf_path)

            logger.info("PDF export - OK: %s", pdf_path)
            return (True, pdf_path)

        except Exception as e:
            logger.error("PDF export - error COM: %s", e)
            return (False, f"Error al exportar: {e}")
        finally:
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel:
                try:
                    excel.Quit()
                except Exception:
                    pass
            if com_initialized:
                try:
                    import pythoncom as _pc
                    _pc.CoUninitialize()
                except Exception:
                    pass

