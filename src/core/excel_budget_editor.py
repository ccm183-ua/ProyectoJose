"""
Edición de presupuestos Excel mediante openpyxl: carga, alta/baja/modificación
de filas y recálculo de totales.
"""

import logging
import os

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Filas de la plantilla (sección de datos y totales en openpyxl)
DATA_START_ROW = 12
SUBTOTAL_ROW = 15
IVA_ROW = 16
TOTAL_ROW = 17


class BudgetEditor:
    """CRUD de filas de presupuesto y recálculo de totales en Excel."""

    def load_budget(self, file_path):
        """
        Carga un presupuesto desde un archivo Excel.

        Args:
            file_path: Ruta del archivo Excel

        Returns:
            Workbook: Objeto Workbook o None si hay error.
                      El llamante debe cerrar el workbook con wb.close().
        """
        try:
            if not os.path.exists(file_path):
                return None

            wb = load_workbook(file_path)
            return wb
        except Exception:
            logger.debug("No se pudo cargar el workbook: %s", file_path)
            return None

    def add_budget_row(self, file_path, budget_row):
        """
        Añade una fila de presupuesto al archivo Excel.

        Args:
            file_path: Ruta del archivo Excel
            budget_row: Diccionario con los datos de la fila

        Returns:
            bool: True si se añadió correctamente, False en caso contrario
        """
        wb = None
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            start_row = DATA_START_ROW
            last_data_row = start_row - 1
            total_row_start = None

            for row_idx in range(start_row, min(ws.max_row + 10, 30)):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value:
                    cell_str = str(cell_value).upper()
                    if ('SUBTOTAL' in cell_str or 'IVA' in cell_str or 'TOTAL' in cell_str) and total_row_start is None:
                        total_row_start = row_idx
                        break
                if cell_value is not None and cell_value != '':
                    last_data_row = row_idx

            if total_row_start is None:
                total_row_start = 15

            if last_data_row < start_row:
                ws.insert_rows(start_row)
                empty_row = start_row
            else:
                empty_row = last_data_row + 1
                if empty_row < total_row_start:
                    ws.insert_rows(empty_row)
                else:
                    ws.insert_rows(total_row_start)
                    empty_row = total_row_start

            ws.cell(row=empty_row, column=1).value = budget_row.get('concepto', '')
            ws.cell(row=empty_row, column=2).value = budget_row.get('cantidad', 0)
            ws.cell(row=empty_row, column=3).value = budget_row.get('unidad', '')
            ws.cell(row=empty_row, column=4).value = budget_row.get('precio_unitario', 0)
            ws.cell(row=empty_row, column=5).value = budget_row.get('importe', 0)

            wb.save(file_path)
            wb.close()
            wb = None

            self.recalculate_totals(file_path)

            return True

        except Exception as e:
            logger.exception("Error al añadir fila")
            return False
        finally:
            if wb is not None:
                wb.close()

    def modify_budget_row(self, file_path, row_index, new_data):
        """
        Modifica una fila de presupuesto.

        Args:
            file_path: Ruta del archivo Excel
            row_index: Índice de la fila a modificar
            new_data: Diccionario con los nuevos datos

        Returns:
            bool: True si se modificó correctamente, False en caso contrario
        """
        wb = None
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            actual_row = 11 + row_index

            if actual_row <= ws.max_row:
                ws.cell(row=actual_row, column=1).value = new_data.get('concepto', '')
                ws.cell(row=actual_row, column=2).value = new_data.get('cantidad', 0)
                ws.cell(row=actual_row, column=3).value = new_data.get('unidad', '')
                ws.cell(row=actual_row, column=4).value = new_data.get('precio_unitario', 0)
                ws.cell(row=actual_row, column=5).value = new_data.get('importe', 0)

                wb.save(file_path)
                wb.close()
                wb = None

                self.recalculate_totals(file_path)
                return True

            return False

        except Exception as e:
            logger.exception("Error al modificar fila")
            return False
        finally:
            if wb is not None:
                wb.close()

    def delete_budget_row(self, file_path, row_index):
        """
        Elimina una fila de presupuesto.

        Args:
            file_path: Ruta del archivo Excel
            row_index: Índice de la fila a eliminar

        Returns:
            bool: True si se eliminó correctamente, False en caso contrario
        """
        wb = None
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            actual_row = 11 + row_index

            if actual_row <= ws.max_row:
                ws.delete_rows(actual_row)

                wb.save(file_path)
                wb.close()
                wb = None

                self.recalculate_totals(file_path)
                return True

            return False

        except Exception as e:
            logger.exception("Error al eliminar fila")
            return False
        finally:
            if wb is not None:
                wb.close()

    def recalculate_totals(self, file_path):
        """
        Recalcula los totales del presupuesto.

        Args:
            file_path: Ruta del archivo Excel

        Returns:
            bool: True si se recalculó correctamente, False en caso contrario
        """
        wb = None
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            start_row = DATA_START_ROW
            end_row = start_row - 1

            for row_idx in range(start_row, SUBTOTAL_ROW):
                if ws.cell(row=row_idx, column=1).value is not None and ws.cell(row=row_idx, column=1).value != '':
                    end_row = row_idx

            if end_row < start_row:
                for row_idx in range(start_row, ws.max_row + 1):
                    cell_value = str(ws.cell(row=row_idx, column=1).value or '').upper()
                    if 'SUBTOTAL' in cell_value or 'IVA' in cell_value or 'TOTAL' in cell_value:
                        break
                    if ws.cell(row=row_idx, column=1).value is not None and ws.cell(row=row_idx, column=1).value != '':
                        end_row = row_idx

            if end_row < start_row:
                end_row = min(ws.max_row, SUBTOTAL_ROW - 1)

            if end_row >= start_row:
                subtotal_formula = f"=SUM(E{start_row}:E{end_row})"
                ws[f'E{SUBTOTAL_ROW}'] = subtotal_formula
                ws[f'E{SUBTOTAL_ROW}'].number_format = '#,##0.00 €'

            ws[f'E{IVA_ROW}'] = f'=E{SUBTOTAL_ROW}*0.21'
            ws[f'E{IVA_ROW}'].number_format = '#,##0.00 €'

            ws[f'E{TOTAL_ROW}'] = f'=E{SUBTOTAL_ROW}+E{IVA_ROW}'
            ws[f'E{TOTAL_ROW}'].number_format = '#,##0.00 €'

            wb.save(file_path)
            return True

        except Exception as e:
            logger.exception("Error al recalcular totales")
            return False
        finally:
            if wb is not None:
                wb.close()

    def save_budget(self, file_path):
        """
        Guarda un presupuesto.

        Args:
            file_path: Ruta del archivo Excel

        Returns:
            bool: True si se guardó correctamente, False en caso contrario
        """
        try:
            return os.path.exists(file_path)
        except Exception:
            logger.debug("Error al verificar archivo: %s", file_path)
            return False
