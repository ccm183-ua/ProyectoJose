"""
Lector del Excel de relación de presupuestos.

Lee el fichero maestro donde se listan todos los presupuestos emitidos
y devuelve los datos en el mismo formato que ``ProjectParser``.
"""

from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# Columnas esperadas (A–J) en el mismo orden que el TSV del portapapeles.
_COL_MAP = {
    0: "numero",          # A
    1: "fecha",           # B
    2: "cliente",         # C
    3: "mediacion",       # D
    4: "calle",           # E
    5: "num_calle",       # F
    6: "codigo_postal",   # G
    7: "localidad",       # H
    8: "tipo",            # I
    9: "importe",         # J
}

_HEADER_MARKER = "Nº"


class ExcelRelationReader:
    """Lee un Excel de relación de presupuestos y devuelve una lista de dicts."""

    def read(self, file_path: str) -> Tuple[List[Dict], Optional[str]]:
        """
        Lee todas las filas de presupuestos del fichero indicado.

        Returns:
            ``(lista_de_dicts, None)`` en caso de éxito, o
            ``([], mensaje_error)`` si ocurre algún problema.
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except FileNotFoundError:
            return [], f"Archivo no encontrado: {file_path}"
        except Exception as exc:
            return [], f"No se pudo abrir el archivo: {exc}"

        try:
            ws = wb.active
            if ws is None:
                return [], "El archivo no contiene hojas"

            header_row = self._find_header_row(ws)
            if header_row is None:
                return [], "No se encontraron las cabeceras (columna A con 'Nº')"

            budgets: List[Dict] = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                parsed = self._parse_row(row)
                if parsed is not None:
                    budgets.append(parsed)

            return budgets, None
        finally:
            wb.close()

    # ── Helpers ───────────────────────────────────────────────────────

    @staticmethod
    def _find_header_row(ws) -> Optional[int]:
        """Busca la fila cuya columna A contenga 'Nº'."""
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=1, values_only=True), start=1):
            cell = row[0]
            if cell is not None and str(cell).strip() == _HEADER_MARKER:
                return row_idx
        return None

    @staticmethod
    def _format_date(value) -> str:
        """Convierte un valor de celda a formato DD-MM-YY."""
        if isinstance(value, datetime):
            return value.strftime("%d-%m-%y")
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        # Si ya viene como DD-MM-YY lo dejamos tal cual
        return text

    @classmethod
    def _parse_row(cls, row: tuple) -> Optional[Dict]:
        """Convierte una fila en un dict o devuelve ``None`` si no es válida."""
        if not row or len(row) < 1:
            return None

        raw_numero = row[0]
        if raw_numero is None:
            return None
        numero_str = str(raw_numero).strip()
        if not numero_str:
            return None

        def _cell(idx: int) -> str:
            if idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        fecha = cls._format_date(row[1] if len(row) > 1 else None)

        importe_raw = row[9] if len(row) > 9 else None
        importe = ""
        if importe_raw is not None:
            try:
                importe = str(float(importe_raw))
            except (ValueError, TypeError):
                importe = str(importe_raw).strip()

        return {
            "numero": numero_str,
            "fecha": fecha,
            "cliente": _cell(2),
            "mediacion": _cell(3),
            "calle": _cell(4),
            "num_calle": _cell(5),
            "codigo_postal": _cell(6),
            "localidad": _cell(7),
            "tipo": _cell(8),
            "importe": importe,
        }
