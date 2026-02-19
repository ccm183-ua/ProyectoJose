"""
Gestor de archivos Excel.
"""

import os
import re
import shutil
import tempfile
import zipfile
from xml.sax.saxutils import escape as xml_escape

from openpyxl import load_workbook
from src.core.template_manager import TemplateManager

# Hoja de datos en la plantilla 122-20
SHEET_12220 = "xl/worksheets/sheet1.xml"


def _replace_cell_in_sheet_xml(sheet_xml, ref, value):
    """Sustituye el valor de la celda ref en el XML de la hoja; conserva el estilo (s=...)."""
    escaped = xml_escape(str(value))
    # Celda: <c r="E5" s="64"/> o <c r="E5" s="64">...</c>; conservar estilo
    pattern = r'<c r="' + re.escape(ref) + r'" ([^>]*?)(?:/>|>.*?</c>)'
    match = re.search(pattern, sheet_xml, re.DOTALL)
    if not match:
        return sheet_xml
    attrs = match.group(1)
    style = re.search(r's="\d+"', attrs)
    style_str = (" " + style.group(0)) if style else ""
    new_cell = f'<c r="{ref}"{style_str} t="inlineStr"><is><t>{escaped}</t></is></c>'
    return re.sub(pattern, new_cell, sheet_xml, count=1, flags=re.DOTALL)


_UNIDADES = (
    '', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO',
    'NUEVE', 'DIEZ', 'ONCE', 'DOCE', 'TRECE', 'CATORCE', 'QUINCE',
    'DIECISÉIS', 'DIECISIETE', 'DIECIOCHO', 'DIECINUEVE', 'VEINTE',
    'VEINTIÚN', 'VEINTIDÓS', 'VEINTITRÉS', 'VEINTICUATRO', 'VEINTICINCO',
    'VEINTISÉIS', 'VEINTISIETE', 'VEINTIOCHO', 'VEINTINUEVE',
)
_DECENAS = (
    '', '', '', 'TREINTA', 'CUARENTA', 'CINCUENTA', 'SESENTA',
    'SETENTA', 'OCHENTA', 'NOVENTA',
)
_CENTENAS = (
    '', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS',
    'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS',
    'NOVECIENTOS',
)


def _numero_a_letras(n: int) -> str:
    """Convierte un entero (0 – 999 999 999) a texto en español (mayúsculas)."""
    if n == 0:
        return 'CERO'
    if n == 100:
        return 'CIEN'

    partes = []

    if n >= 1_000_000:
        millones = n // 1_000_000
        n %= 1_000_000
        if millones == 1:
            partes.append('UN MILLÓN')
        else:
            partes.append(f'{_numero_a_letras(millones)} MILLONES')

    if n >= 1000:
        miles = n // 1000
        n %= 1000
        if miles == 1:
            partes.append('MIL')
        else:
            partes.append(f'{_numero_a_letras(miles)} MIL')

    if n >= 100:
        if n == 100:
            partes.append('CIEN')
            return ' '.join(partes)
        partes.append(_CENTENAS[n // 100])
        n %= 100

    if 0 < n < 30:
        partes.append(_UNIDADES[n])
    elif n >= 30:
        decena = _DECENAS[n // 10]
        unidad = _UNIDADES[n % 10]
        partes.append(f'{decena} Y {unidad}' if unidad else decena)

    return ' '.join(p for p in partes if p)


def _euros_en_letras(importe: float) -> str:
    """Devuelve *importe* en formato 'CIENTO VEINTITRÉS EUROS CON CUARENTA Y CINCO CÉNTIMOS'."""
    centimos_total = round(importe * 100)
    euros = centimos_total // 100
    centimos = centimos_total % 100

    txt = _numero_a_letras(euros)
    txt += ' EURO' if euros == 1 else ' EUROS'

    if centimos:
        txt += f' CON {_numero_a_letras(centimos)}'
        txt += ' CÉNTIMO' if centimos == 1 else ' CÉNTIMOS'

    return txt


class ExcelManager:
    """Clase para gestionar archivos Excel."""
    
    def __init__(self):
        """Inicializa el gestor de Excel."""
        self.template_manager = TemplateManager()
    
    def create_from_template(self, template_path, output_path, data):
        """
        Crea un archivo Excel desde una plantilla rellenando los datos.
        
        Args:
            template_path: Ruta de la plantilla
            output_path: Ruta donde guardar el archivo
            data: Diccionario con los datos a rellenar
            
        Returns:
            bool: True si se creó correctamente, False en caso contrario
        """
        try:
            # Crear directorio de salida primero
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)

            # Copiar plantilla al destino (logo, imágenes y formato quedan intactos)
            shutil.copy2(template_path, output_path)

            # Preparar datos del proyecto
            nombre_obra = data.get('nombre_obra', '')
            if not nombre_obra:
                nombre_obra = f"{data.get('direccion', '')} {data.get('numero', '')}".strip()
            direccion_parts = []
            if data.get('calle'):
                direccion_parts.append(data.get('calle'))
            if data.get('num_calle'):
                direccion_parts.append(f"Nº {data.get('num_calle')}")
            direccion_solo_calle_numero = ' '.join(direccion_parts)
            if not direccion_solo_calle_numero:
                direccion_solo_calle_numero = (data.get('direccion', '') or '').strip()
                if data.get('numero'):
                    direccion_solo_calle_numero = f"{direccion_solo_calle_numero} Nº {data.get('numero')}".strip()

            # Rellenar solo las celdas de datos en el XML de la hoja (sin abrir con openpyxl = logo/título intactos)
            self._patch_sheet2_cells_12220(output_path, data, nombre_obra, direccion_solo_calle_numero)
            return True
            
        except Exception as e:
            print(f"Error al crear archivo Excel: {e}")
            return False

    def _patch_sheet2_cells_12220(self, output_path, data, nombre_obra, direccion_solo_calle_numero):
        """
        Modifica solo el XML de la hoja de datos (sheet2) dentro del xlsx, sin tocar
        medios (logo, imágenes) ni dibujos. Así el logo y el título quedan intactos.
        """
        # Valores para cada celda (inlineStr para no tocar sharedStrings)
        fecha = data.get('fecha', '')
        if fecha:
            try:
                parts = str(fecha).strip().split('-')
                if len(parts) == 3:
                    fecha = f"{parts[0]}/{parts[1]}/{parts[2]}"
            except Exception:
                pass
        obra_texto = (data.get('tipo', '') or nombre_obra or '').strip()
        obra_final = f"Obra: {obra_texto}." if obra_texto else "Obra:"

        numero_pres = str(data.get('numero_proyecto', '') or data.get('numero', '') or '').strip()
        fecha_raw = data.get('fecha', '') or ''
        year_suffix = ''
        try:
            fecha_parts = str(fecha_raw).strip().split('-')
            if len(fecha_parts) == 3:
                year_suffix = fecha_parts[2]
        except Exception:
            pass
        if numero_pres and year_suffix:
            numero_pres = f"{numero_pres}/{year_suffix}"

        cliente = (data.get('cliente', '') or '').strip()

        celdas = {
            "E5": numero_pres,
            "H5": fecha or '',
            "B7": cliente,
            "H7": (data.get('admin_cif', '') or '').strip(),
            "B9": (direccion_solo_calle_numero or '').strip(),
            "H9": str(data.get('codigo_postal', '') or '').strip(),
            "B11": (data.get('admin_email', '') or '').strip(),
            "H11": (data.get('admin_telefono', '') or '').strip(),
            "A14": obra_final,
            "A57": cliente,
        }

        with zipfile.ZipFile(output_path, "r") as z_in:
            namelist = z_in.namelist()
            sheet_content = z_in.read(SHEET_12220).decode("utf-8")
            otros = {n: z_in.read(n) for n in namelist if n != SHEET_12220}

        for ref, valor in celdas.items():
            sheet_content = _replace_cell_in_sheet_xml(sheet_content, ref, valor)

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        try:
            os.close(fd)
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z_out:
                for name in namelist:
                    if name == SHEET_12220:
                        z_out.writestr(name, sheet_content.encode("utf-8"))
                    else:
                        z_out.writestr(name, otros[name])
            shutil.move(tmp_path, output_path)
        except Exception:
            if os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
            raise

    def insert_partidas_via_xml(self, file_path, partidas):
        """
        Inserta partidas en el Excel usando manipulación XML directa en sheet2.

        Reemplaza las partidas de ejemplo de la plantilla (filas 17-26) con las
        partidas proporcionadas, manteniendo estilos y formato del template.

        Estructura de columnas en la plantilla:
            A = Número (1.1, 1.2...)
            B = Unidad (m2, ud, ml...)
            C-F = Descripción (celdas C:F combinadas)
            G = Cantidad
            H = Precio unitario
            I = Total (fórmula =G*H)

        Args:
            file_path: Ruta del archivo Excel ya creado desde plantilla.
            partidas: Lista de dicts con concepto, cantidad, unidad, precio_unitario.

        Returns:
            bool: True si se insertaron correctamente.
        """
        if not partidas:
            return True  # Nada que insertar

        try:
            with zipfile.ZipFile(file_path, "r") as z_in:
                namelist = z_in.namelist()
                sheet_content = z_in.read(SHEET_12220).decode("utf-8")
                otros = {n: z_in.read(n) for n in namelist if n != SHEET_12220}

            sheet_content = self._replace_partidas_in_xml(sheet_content, partidas)

            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            try:
                os.close(fd)
                with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z_out:
                    for name in namelist:
                        if name == SHEET_12220:
                            z_out.writestr(name, sheet_content.encode("utf-8"))
                        elif name == "xl/workbook.xml":
                            wb_xml = otros[name].decode("utf-8")
                            wb_xml = re.sub(
                                r'<calcPr [^/]*/>', 
                                '<calcPr calcId="144525" fullCalcOnLoad="1"/>',
                                wb_xml,
                            )
                            z_out.writestr(name, wb_xml.encode("utf-8"))
                        else:
                            z_out.writestr(name, otros[name])
                shutil.move(tmp_path, file_path)
            except Exception:
                if os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
                raise

            return True

        except Exception as e:
            print(f"Error al insertar partidas: {e}")
            return False

    @staticmethod
    def _estimate_row_height(titulo, descripcion, chars_per_line=55, line_height=14.5):
        """
        Estima la altura de fila necesaria para el texto de una partida.

        Calcula el número de líneas que ocupará el texto en la celda combinada
        C:F (ancho aprox. ~55 caracteres en Calibri 10pt) y devuelve la altura
        en puntos Excel.

        Args:
            titulo: Texto del título (1 línea).
            descripcion: Texto de la descripción (puede ocupar varias líneas).
            chars_per_line: Caracteres aproximados que caben por línea.
            line_height: Altura en puntos por línea de texto.

        Returns:
            Altura de fila como string (en puntos).
        """
        # Línea 1: título (siempre 1 línea)
        lines = 1

        # Líneas de descripción
        if descripcion:
            desc_len = len(descripcion)
            desc_lines = max(1, -(-desc_len // chars_per_line))  # Redondeo hacia arriba
            lines += desc_lines

        # Altura = líneas * alto_por_línea + padding superior e inferior
        height = lines * line_height + 8

        # Mínimo 30, máximo razonable 200
        height = max(30, min(200, height))

        return str(round(height, 1))

    def _replace_partidas_in_xml(self, sheet_xml, partidas):
        """
        Reemplaza las filas de partidas de ejemplo (17-26) con las partidas reales.

        La plantilla tiene filas de datos en 17, 19, 21, 23, 25 (impares)
        y filas separadoras vacías en 18, 20, 22, 24, 26 (pares).
        Fila 27 tiene el subtotal con =SUM(I17:I26).

        Este método:
        1. Elimina todas las filas entre 17 y 26 (datos + separadores de ejemplo)
        2. Genera nuevas filas de datos y separadores para cada partida
        3. Genera una fila de subtotal con la fórmula SUM correcta
        4. Renumera las filas posteriores (27+) para que encajen

        Estilos de la plantilla sheet1 (atributo s= de cada celda):
        - Fila de datos: A=s31, B=s31, C=s32(merged C:F), D=s32, E=s32, F=s33, G=s34, H=s35, I=s35
        - Fila separadora: A=s31, B=s31, C=s36, D=s32, E=s32, F=s33, G=s34, H=s35, I=s35
        - Fila subtotal: A=s39, B=s40, C=s41(merged C:G), D-G=s42, H=s43, I=s54

        Args:
            sheet_xml: Contenido XML de sheet2.
            partidas: Lista de partidas a insertar.

        Returns:
            XML modificado con las partidas insertadas.
        """
        # --- Paso 1: Identificar la zona de partidas (filas 17-26) y eliminarla ---
        # Eliminar filas 17 a 26 del XML
        for row_num in range(17, 27):
            pattern = r'<row r="' + str(row_num) + r'"[^>]*>.*?</row>'
            sheet_xml = re.sub(pattern, '', sheet_xml, flags=re.DOTALL)

        # --- Paso 2: Generar las nuevas filas de datos + separadores ---
        new_rows_xml = []
        first_data_row = 17
        current_row = first_data_row

        for idx, partida in enumerate(partidas):
            num = f"1.{idx + 1}"
            unidad = xml_escape(str(partida.get('unidad', 'ud')))
            cantidad = partida.get('cantidad', 1)
            precio = partida.get('precio_unitario', 0)

            # Asegurar valores numéricos
            try:
                cantidad = float(cantidad)
            except (ValueError, TypeError):
                cantidad = 1.0
            try:
                precio = float(precio)
            except (ValueError, TypeError):
                precio = 0.0

            # Construir celda C con rich text: título en negrita + descripción normal
            titulo = xml_escape(str(partida.get('titulo', '')))
            descripcion = xml_escape(str(partida.get('descripcion', '')))

            if titulo and descripcion:
                celda_c = (
                    f'<c r="C{current_row}" s="32" t="inlineStr"><is>'
                    f'<r><rPr><b/><sz val="10"/><rFont val="Calibri"/></rPr>'
                    f'<t>{titulo}</t></r>'
                    f'<r><rPr><sz val="10"/><rFont val="Calibri"/></rPr>'
                    f'<t xml:space="preserve">&#10;{descripcion}</t></r>'
                    f'</is></c>'
                )
                row_height = self._estimate_row_height(titulo, descripcion)
            elif titulo:
                celda_c = (
                    f'<c r="C{current_row}" s="32" t="inlineStr"><is>'
                    f'<r><rPr><b/><sz val="10"/><rFont val="Calibri"/></rPr>'
                    f'<t>{titulo}</t></r>'
                    f'</is></c>'
                )
                row_height = self._estimate_row_height(titulo, '')
            else:
                concepto = xml_escape(str(partida.get('concepto', '')))
                celda_c = (
                    f'<c r="C{current_row}" s="32" t="inlineStr">'
                    f'<is><t>{concepto}</t></is></c>'
                )
                row_height = self._estimate_row_height(concepto, '')

            total = round(cantidad * precio, 2)

            data_row = (
                f'<row r="{current_row}" spans="1:9" ht="{row_height}" customHeight="1">'
                f'<c r="A{current_row}" s="31" t="inlineStr"><is><t>{num}</t></is></c>'
                f'<c r="B{current_row}" s="31" t="inlineStr"><is><t>{unidad}</t></is></c>'
                f'{celda_c}'
                f'<c r="D{current_row}" s="32"/>'
                f'<c r="E{current_row}" s="32"/>'
                f'<c r="F{current_row}" s="33"/>'
                f'<c r="G{current_row}" s="34"><v>{cantidad}</v></c>'
                f'<c r="H{current_row}" s="35"><v>{precio}</v></c>'
                f'<c r="I{current_row}" s="35"><f>G{current_row}*H{current_row}</f><v>{total}</v></c>'
                f'</row>'
            )
            new_rows_xml.append(data_row)
            current_row += 1

            spacer_row = (
                f'<row r="{current_row}" spans="1:9" customHeight="1">'
                f'<c r="A{current_row}" s="31"/>'
                f'<c r="B{current_row}" s="31"/>'
                f'<c r="C{current_row}" s="36"/>'
                f'<c r="D{current_row}" s="32"/>'
                f'<c r="E{current_row}" s="32"/>'
                f'<c r="F{current_row}" s="33"/>'
                f'<c r="G{current_row}" s="34"/>'
                f'<c r="H{current_row}" s="35"/>'
                f'<c r="I{current_row}" s="35"/>'
                f'</row>'
            )
            new_rows_xml.append(spacer_row)
            current_row += 1

        last_data_row = current_row - 1  # Última fila (separador de la última partida)

        # --- Paso 3: Fila de subtotal ---
        subtotal_row_num = current_row
        # Rango de la SUM: desde la primera fila de datos hasta la última fila antes del subtotal
        subtotal_label = xml_escape("Total presupuesto parcial nº 1 ACTUACIONES.")
        grand_total = sum(
            round(float(p.get('cantidad', 1)) * float(p.get('precio_unitario', 0)), 2)
            for p in partidas
        )
        subtotal_row = (
            f'<row r="{subtotal_row_num}" spans="1:9" customHeight="1">'
            f'<c r="A{subtotal_row_num}" s="39"/>'
            f'<c r="B{subtotal_row_num}" s="40"/>'
            f'<c r="C{subtotal_row_num}" s="41" t="inlineStr"><is><t>{subtotal_label}</t></is></c>'
            f'<c r="D{subtotal_row_num}" s="42"/>'
            f'<c r="E{subtotal_row_num}" s="42"/>'
            f'<c r="F{subtotal_row_num}" s="42"/>'
            f'<c r="G{subtotal_row_num}" s="42"/>'
            f'<c r="H{subtotal_row_num}" s="43"/>'
            f'<c r="I{subtotal_row_num}" s="54">'
            f'<f>SUM(I{first_data_row}:I{last_data_row})</f>'
            f'<v>{grand_total}</v></c>'
            f'</row>'
        )
        new_rows_xml.append(subtotal_row)

        # --- Paso 4: Eliminar fila 27 original (subtotal viejo) ---
        pattern_27 = r'<row r="27"[^>]*>.*?</row>'
        sheet_xml = re.sub(pattern_27, '', sheet_xml, flags=re.DOTALL)

        # --- Paso 5: Renumerar filas 28+ para que sigan después del nuevo subtotal ---
        # Las filas originales 28-57 deben desplazarse según la diferencia
        # Original: subtotal en fila 27, ahora en subtotal_row_num
        offset = subtotal_row_num - 27  # Cuántas filas se desplaza todo

        if offset != 0:
            sheet_xml = self._renumber_rows(sheet_xml, start_from=28, offset=offset)

        # --- Paso 6: Actualizar mergeCells para la zona de datos ---
        # Eliminar merges de las filas de ejemplo (C17:F17, C19:F19, etc. y C27:G27)
        for row_num in range(17, 28):
            sheet_xml = re.sub(
                r'<mergeCell ref="[A-Z]+' + str(row_num) + r':[A-Z]+' + str(row_num) + r'"/>',
                '', sheet_xml
            )

        # Añadir nuevos merges para las filas de datos (C:F combinadas)
        new_merges = []
        row = first_data_row
        for idx in range(len(partidas)):
            new_merges.append(f'<mergeCell ref="C{row}:F{row}"/>')
            row += 2  # Saltar separador

        # Merge para la fila de subtotal (C:G)
        new_merges.append(f'<mergeCell ref="C{subtotal_row_num}:G{subtotal_row_num}"/>')

        # Actualizar merges de filas renumeradas (39+offset, etc.)
        if offset != 0:
            sheet_xml = self._renumber_merges(sheet_xml, start_from=28, offset=offset)

        # Insertar nuevos merges
        merge_insert = ''.join(new_merges)
        sheet_xml = sheet_xml.replace('</mergeCells>', merge_insert + '</mergeCells>')

        # --- Paso 7: Actualizar TODAS las fórmulas de la zona de resumen ---
        r43 = 43 + offset  # "1.- ACTUACIONES." → subtotal
        r45 = 45 + offset  # "Total presupuesto"
        r46 = 46 + offset  # "10% I.V.A."
        r47 = 47 + offset  # "TOTAL PRESUPUESTO, I.V.A. INCLUIDO."
        r49 = 49 + offset  # Texto del importe en letras

        iva_rate = 0.10
        total_sin_iva = grand_total
        iva_amount = round(total_sin_iva * iva_rate, 2)
        total_con_iva = round(total_sin_iva + iva_amount, 2)

        sheet_xml = self._update_formula_ref(
            sheet_xml, r43, 'I', f'I{subtotal_row_num}', total_sin_iva)
        sheet_xml = self._update_formula_ref(
            sheet_xml, r45, 'I', f'SUM(I{r43}:I{r43 + 1})', total_sin_iva)
        sheet_xml = self._update_formula_ref(
            sheet_xml, r46, 'I', f'I{r45}*0.1', iva_amount)
        sheet_xml = self._update_formula_ref(
            sheet_xml, r47, 'I', f'I{r45}+I{r46}', total_con_iva)

        # Actualizar texto del importe en letras (fila 49)
        texto_importe = (
            "Asciende el presupuesto de ejecución material a la expresada "
            f"cantidad de {_euros_en_letras(total_con_iva)} IVA INCLUIDO."
        )
        sheet_xml = self._replace_cell_text(sheet_xml, f'A{r49}', texto_importe, style='47')

        # --- Paso 8: Insertar las nuevas filas en el XML ---
        # Insertamos justo antes de la primera fila que quede después del subtotal
        next_original_row = 28 + offset
        insert_point_pattern = r'(<row r="' + str(next_original_row) + r'")'
        insert_point = re.search(insert_point_pattern, sheet_xml)

        all_new_rows = '\n'.join(new_rows_xml)
        if insert_point:
            sheet_xml = sheet_xml[:insert_point.start()] + all_new_rows + '\n' + sheet_xml[insert_point.start():]
        else:
            # Si no encontramos la fila siguiente, insertar antes de </sheetData>
            sheet_xml = sheet_xml.replace('</sheetData>', all_new_rows + '\n</sheetData>')

        # --- Paso 9: Actualizar dimension ---
        last_row = 57 + offset
        sheet_xml = re.sub(
            r'<dimension ref="[^"]+"/>',
            f'<dimension ref="A1:R{last_row}"/>',
            sheet_xml
        )

        return sheet_xml

    def _renumber_rows(self, sheet_xml, start_from, offset):
        """Renumera filas y sus referencias de celda desde start_from aplicando offset."""
        if offset == 0:
            return sheet_xml

        # Procesar de mayor a menor para evitar conflictos
        # Encontrar todas las filas con r >= start_from
        rows_found = re.findall(r'<row r="(\d+)"', sheet_xml)
        rows_to_renumber = sorted(
            [int(r) for r in rows_found if int(r) >= start_from],
            reverse=True
        )

        for old_row in rows_to_renumber:
            new_row = old_row + offset

            # Renumerar el atributo r de la fila
            sheet_xml = sheet_xml.replace(
                f'<row r="{old_row}"',
                f'<row r="{new_row}"'
            )

            # Renumerar referencias de celdas dentro de esa fila (A28 → A28+offset, etc.)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                        'M', 'N', 'O', 'P', 'Q', 'R']:
                sheet_xml = sheet_xml.replace(
                    f'r="{col}{old_row}"',
                    f'r="{col}{new_row}"'
                )

            # Renumerar fórmulas que referencian filas originales
            # Ej: =I27 → =I{27+offset}, =SUM(I43:I44) → =SUM(I{43+o}:I{44+o})
            # Solo dentro de <f>...</f> tags - esto es complejo, hacerlo caso por caso

        return sheet_xml

    def _renumber_merges(self, sheet_xml, start_from, offset):
        """Renumera mergeCell refs que involucran filas >= start_from."""
        if offset == 0:
            return sheet_xml

        def _replace_merge(match):
            ref = match.group(1)
            # Parse "A39:I39" → cols + rows
            parts = ref.split(':')
            if len(parts) != 2:
                return match.group(0)
            start_ref, end_ref = parts
            start_col = re.match(r'([A-Z]+)', start_ref).group(1)
            start_row = int(re.search(r'(\d+)', start_ref).group(1))
            end_col = re.match(r'([A-Z]+)', end_ref).group(1)
            end_row = int(re.search(r'(\d+)', end_ref).group(1))

            if start_row >= start_from:
                start_row += offset
                end_row += offset
                return f'<mergeCell ref="{start_col}{start_row}:{end_col}{end_row}"/>'
            return match.group(0)

        sheet_xml = re.sub(r'<mergeCell ref="([^"]+)"/>', _replace_merge, sheet_xml)
        return sheet_xml

    def _update_formula_ref(self, sheet_xml, row, col, new_formula, cached_value=None):
        """Actualiza la fórmula (y valor cacheado) de una celda específica."""
        pattern = (
            r'(<c r="' + col + str(row) + r'"[^>]*>)'
            r'(?:<f[^>]*>[^<]*</f>)?'
            r'(?:<v>[^<]*</v>)?'
        )
        match = re.search(pattern, sheet_xml)
        if match:
            v_part = ''
            if cached_value is not None:
                v_part = f'<v>{cached_value}</v>'
            replacement = match.group(1) + f'<f>{new_formula}</f>' + v_part
            sheet_xml = sheet_xml[:match.start()] + replacement + sheet_xml[match.end():]
        return sheet_xml

    def _replace_cell_text(self, sheet_xml, ref, text, style=None):
        """Reemplaza el contenido de texto de una celda (inline string)."""
        escaped = xml_escape(str(text))
        pattern = r'<c r="' + re.escape(ref) + r'"[^>]*?>(?:.*?</c>|/>)'
        match = re.search(pattern, sheet_xml, re.DOTALL)
        if match:
            s_attr = f' s="{style}"' if style else ''
            new_cell = (
                f'<c r="{ref}"{s_attr} t="inlineStr">'
                f'<is><t>{escaped}</t></is></c>'
            )
            sheet_xml = sheet_xml[:match.start()] + new_cell + sheet_xml[match.end():]
        return sheet_xml

    def load_budget(self, file_path):
        """
        Carga un presupuesto desde un archivo Excel.
        
        Args:
            file_path: Ruta del archivo Excel
            
        Returns:
            Workbook: Objeto Workbook o None si hay error
        """
        try:
            if not os.path.exists(file_path):
                return None
            
            wb = load_workbook(file_path)
            return wb
        except Exception as e:
            return None
    
    def add_budget_row(self, file_path, budget_row):
        """
        Añade una fila de presupuesto al archivo Excel.
        
        Args:
            file_path: Ruta del archivo Excel
            budget_row: Diccionario con los datos de la fila
            
        Returns:
            bool: True si se añadió correctamente, False en caso contrario
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar la última fila con datos antes de los totales (filas 15-17)
            # Asumimos que los encabezados están en la fila 11
            start_row = 12
            last_data_row = start_row - 1
            total_row_start = None
            
            # Buscar la última fila con datos y dónde empiezan los totales
            for row_idx in range(start_row, min(ws.max_row + 10, 30)):  # Buscar hasta la fila 30
                cell_value = ws.cell(row=row_idx, column=1).value
                # Detener si encontramos "SUBTOTAL", "IVA" o "TOTAL"
                if cell_value:
                    cell_str = str(cell_value).upper()
                    if ('SUBTOTAL' in cell_str or 'IVA' in cell_str or 'TOTAL' in cell_str) and total_row_start is None:
                        total_row_start = row_idx
                        break
                if cell_value is not None and cell_value != '':
                    last_data_row = row_idx
            
            # Si no encontramos dónde empiezan los totales, asumimos fila 15
            if total_row_start is None:
                total_row_start = 15
            
            # Siempre insertar una nueva fila después de la última fila con datos
            if last_data_row < start_row:
                # No hay datos, insertar en start_row
                ws.insert_rows(start_row)
                empty_row = start_row
            else:
                # Insertar después de la última fila con datos, pero antes de los totales
                empty_row = last_data_row + 1
                # Solo insertar si no estamos en la fila de totales
                if empty_row < total_row_start:
                    ws.insert_rows(empty_row)
                else:
                    # Si ya estamos en o después de los totales, insertar antes de ellos
                    ws.insert_rows(total_row_start)
                    empty_row = total_row_start
            
            # Añadir datos
            ws.cell(row=empty_row, column=1).value = budget_row.get('concepto', '')
            ws.cell(row=empty_row, column=2).value = budget_row.get('cantidad', 0)
            ws.cell(row=empty_row, column=3).value = budget_row.get('unidad', '')
            ws.cell(row=empty_row, column=4).value = budget_row.get('precio_unitario', 0)
            ws.cell(row=empty_row, column=5).value = budget_row.get('importe', 0)
            
            # Guardar el archivo
            wb.save(file_path)
            
            # Actualizar fórmulas de totales (esto recargará el archivo)
            self.recalculate_totals(file_path)
            
            return True
            
        except Exception as e:
            print(f"Error al añadir fila: {e}")
            return False
    
    def modify_budget_row(self, file_path, row_index, new_data):
        """
        Modifica una fila de presupuesto.
        
        Args:
            file_path: Ruta del archivo Excel
            row_index: Índice de la fila a modificar
            new_data: Diccionario con los nuevos datos
            
        Returns:
            bool: True si se modificó correctamente, False en caso contrario
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Modificar datos (row_index es 1-based, pero ajustamos para la fila real)
            actual_row = 11 + row_index  # Asumiendo que los datos empiezan en fila 12
            
            if actual_row <= ws.max_row:
                ws.cell(row=actual_row, column=1).value = new_data.get('concepto', '')
                ws.cell(row=actual_row, column=2).value = new_data.get('cantidad', 0)
                ws.cell(row=actual_row, column=3).value = new_data.get('unidad', '')
                ws.cell(row=actual_row, column=4).value = new_data.get('precio_unitario', 0)
                ws.cell(row=actual_row, column=5).value = new_data.get('importe', 0)
                
                # Recalcular totales
                self.recalculate_totals(file_path)
                
                wb.save(file_path)
                return True
            
            return False
            
        except Exception as e:
            print(f"Error al modificar fila: {e}")
            return False
    
    def delete_budget_row(self, file_path, row_index):
        """
        Elimina una fila de presupuesto.
        
        Args:
            file_path: Ruta del archivo Excel
            row_index: Índice de la fila a eliminar
            
        Returns:
            bool: True si se eliminó correctamente, False en caso contrario
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            actual_row = 11 + row_index
            
            if actual_row <= ws.max_row:
                ws.delete_rows(actual_row)
                
                # Recalcular totales
                self.recalculate_totals(file_path)
                
                wb.save(file_path)
                return True
            
            return False
            
        except Exception as e:
            print(f"Error al eliminar fila: {e}")
            return False
    
    def recalculate_totals(self, file_path):
        """
        Recalcula los totales del presupuesto.
        
        Args:
            file_path: Ruta del archivo Excel
            
        Returns:
            bool: True si se recalculó correctamente, False en caso contrario
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar fila de inicio de datos (asumimos fila 12)
            start_row = 12
            end_row = start_row - 1
            
            # Buscar última fila con datos antes de los totales (filas 15-17)
            # Buscar desde start_row hasta la fila 14 (antes de los totales)
            for row_idx in range(start_row, 15):
                if ws.cell(row=row_idx, column=1).value is not None and ws.cell(row=row_idx, column=1).value != '':
                    end_row = row_idx
            
            # Si no encontramos datos en las filas 12-14, buscar más abajo
            if end_row < start_row:
                # Buscar en un rango más amplio, pero antes de los totales
                for row_idx in range(start_row, ws.max_row + 1):
                    # Detener si encontramos "SUBTOTAL", "IVA" o "TOTAL"
                    cell_value = str(ws.cell(row=row_idx, column=1).value or '').upper()
                    if 'SUBTOTAL' in cell_value or 'IVA' in cell_value or 'TOTAL' in cell_value:
                        break
                    if ws.cell(row=row_idx, column=1).value is not None and ws.cell(row=row_idx, column=1).value != '':
                        end_row = row_idx
            
            # Si aún no encontramos, usar max_row pero asegurarnos de que no sea una fila de totales
            if end_row < start_row:
                end_row = min(ws.max_row, 14)  # Máximo hasta la fila 14
            
            # Actualizar fórmula de subtotal (asumimos fila 15)
            if end_row >= start_row:
                subtotal_formula = f"=SUM(E{start_row}:E{end_row})"
                ws['E15'] = subtotal_formula
                ws['E15'].number_format = '#,##0.00 €'
            
            # Actualizar fórmula de IVA (asumimos fila 16)
            ws['E16'] = '=E15*0.21'
            ws['E16'].number_format = '#,##0.00 €'
            
            # Actualizar fórmula de TOTAL (asumimos fila 17)
            ws['E17'] = '=E15+E16'
            ws['E17'].number_format = '#,##0.00 €'
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error al recalcular totales: {e}")
            return False
    
    def save_budget(self, file_path):
        """
        Guarda un presupuesto.
        
        Args:
            file_path: Ruta del archivo Excel
            
        Returns:
            bool: True si se guardó correctamente, False en caso contrario
        """
        try:
            # Este método se puede usar para guardar cambios si se está editando
            # Por ahora, simplemente verificamos que el archivo existe
            return os.path.exists(file_path)
        except Exception:
            return False
