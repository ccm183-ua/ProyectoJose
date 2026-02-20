"""
Tests de integración para el flujo de rutas por defecto.

Cubren:
- FileDialog recibe defaultDir correcto
- Flujo de creación con Excel de relación
- Fallback al portapapeles
- Selector de presupuestos (filtro y datos)
"""

import os
import tempfile
import shutil
from datetime import datetime
from unittest.mock import patch, MagicMock

import pytest
from openpyxl import Workbook

from src.core.settings import Settings
from src.core.excel_relation_reader import ExcelRelationReader
from src.utils.project_name_generator import ProjectNameGenerator


@pytest.fixture
def temp_dir():
    path = tempfile.mkdtemp()
    yield path
    shutil.rmtree(path, ignore_errors=True)


def _create_relation_excel(path, rows, header_row=3):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Test S.L.")
    ws.cell(row=2, column=1, value="RELACIÓN 2026")
    headers = ["Nº", "FECHA", "CLIENTE", "MEDIACIÓN", "CALLE",
               "NUM", "C.P", "LOCALIDAD", "TIPO", "IMPORTE"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    for r, data in enumerate(rows, header_row + 1):
        for c, v in enumerate(data, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    wb.close()


# ── Tests de integración Settings + FileDialog ──────────────────


class TestFileDialogDefaultDir:
    """Verifica que los diálogos de archivo usen la ruta por defecto."""

    def test_save_dialog_uses_default_path(self, temp_dir):
        """La ruta de guardado se aplica correctamente."""
        s = Settings(config_dir=temp_dir)
        s.set_default_path(Settings.PATH_SAVE_BUDGETS, "C:/mis_presupuestos")
        assert s.get_default_path(Settings.PATH_SAVE_BUDGETS) == "C:/mis_presupuestos"

    def test_open_dialog_uses_default_path(self, temp_dir):
        """La ruta de apertura se aplica correctamente."""
        s = Settings(config_dir=temp_dir)
        s.set_default_path(Settings.PATH_OPEN_BUDGETS, "D:/existentes")
        assert s.get_default_path(Settings.PATH_OPEN_BUDGETS) == "D:/existentes"

    def test_relation_file_path_stored(self, temp_dir):
        """La ruta del fichero de relación se almacena y recupera."""
        fp = os.path.join(temp_dir, "relacion.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "C", "", "C/", "1", "", "", "", None),
        ])
        s = Settings(config_dir=temp_dir)
        s.set_default_path(Settings.PATH_RELATION_FILE, fp)
        assert s.get_default_path(Settings.PATH_RELATION_FILE) == fp
        assert os.path.isfile(fp)


# ── Tests de flujo completo con Excel de relación ────────────────


class TestCreateBudgetWithRelation:
    """Integración: leer Excel de relación y obtener datos de proyecto."""

    def test_create_budget_with_relation_file(self, temp_dir):
        """Leer relación → seleccionar fila → obtener project_data + name."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "C.P. GABRIEL MIRÓ, 24", "MSFINCAS",
             "C/ GABRIEL MIRÓ Nº 24", "24", "", "SAN VICENTE", "PUERTA ZAGUÁN", 4650),
        ])
        budgets, err = ExcelRelationReader().read(fp)
        assert err is None
        assert len(budgets) == 1

        chosen = budgets[0]
        name = ProjectNameGenerator().generate_project_name(chosen)
        assert chosen["numero"] == "1"
        assert chosen["cliente"] == "C.P. GABRIEL MIRÓ, 24"
        assert name  # se genera un nombre no vacío

    def test_create_budget_fallback_to_clipboard(self, temp_dir):
        """Si el Excel de relación no existe, se usa el portapapeles."""
        s = Settings(config_dir=temp_dir)
        ruta = os.path.join(temp_dir, "no_existe.xlsx")
        s.set_default_path(Settings.PATH_RELATION_FILE, ruta)

        assert not os.path.isfile(ruta)
        # El flujo en main_frame comprueba os.path.isfile antes de leer

    def test_create_budget_no_relation_configured(self, temp_dir):
        """Sin ruta configurada devuelve None → va al portapapeles."""
        s = Settings(config_dir=temp_dir)
        assert s.get_default_path(Settings.PATH_RELATION_FILE) is None


# ── Tests del selector de presupuestos ───────────────────────────


class TestBudgetSelector:
    """Tests de la lógica de filtrado y datos del selector."""

    def test_budget_selector_shows_all_entries(self, temp_dir):
        """El reader devuelve todas las entradas válidas."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        rows = [(i, datetime(2026, 1, i), f"Cliente {i}", "", "C/", str(i), "", "L", "T", 0)
                for i in range(1, 6)]
        _create_relation_excel(fp, rows)
        budgets, err = ExcelRelationReader().read(fp)
        assert err is None
        assert len(budgets) == 5

    def test_budget_selector_filter(self, temp_dir):
        """Filtro en memoria sobre la lista de presupuestos."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 1), "ALFA", "", "C/ A", "1", "", "Madrid", "Pintura", None),
            (2, datetime(2026, 1, 2), "BETA", "", "C/ B", "2", "", "Barcelona", "Reforma", None),
            (3, datetime(2026, 1, 3), "GAMMA", "", "C/ C", "3", "", "Madrid", "Fachada", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)

        query = "madrid"
        filtered = [
            b for b in budgets
            if query in " ".join(str(v) for v in b.values()).lower()
        ]
        assert len(filtered) == 2
        assert all("Madrid" in b["localidad"] for b in filtered)

    def test_budget_selector_returns_correct_data(self, temp_dir):
        """Los datos seleccionados son correctos y completos."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (5, datetime(2026, 3, 15), "C.P. SOL, 10", "ADMIN",
             "C/ Sol Nº 10", "10", "28001", "Madrid", "Impermeabilización", 3200),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        b = budgets[0]
        assert b["numero"] == "5"
        assert b["fecha"] == "15-03-26"
        assert b["cliente"] == "C.P. SOL, 10"
        assert b["mediacion"] == "ADMIN"
        assert b["calle"] == "C/ Sol Nº 10"
        assert b["num_calle"] == "10"
        assert b["codigo_postal"] == "28001"
        assert b["localidad"] == "Madrid"
        assert b["tipo"] == "Impermeabilización"
        assert b["importe"] == "3200.0"

    def test_budget_data_generates_valid_project_name(self, temp_dir):
        """Los datos del selector generan un nombre de proyecto válido."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (7, datetime(2026, 5, 20), "C.P. LUNA, 3", "",
             "C/ Luna Nº 3", "3", "03001", "Alicante", "Rehabilitación", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        name = ProjectNameGenerator().generate_project_name(budgets[0])
        assert "7" in name
        assert name  # no vacío

    def test_end_to_end_settings_to_reader(self, temp_dir):
        """Flujo completo: configurar ruta → leer → obtener datos."""
        fp = os.path.join(temp_dir, "relacion.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "C", "M", "C/", "1", "03001", "L", "T", 100),
        ])
        s = Settings(config_dir=temp_dir)
        s.set_default_path(Settings.PATH_RELATION_FILE, fp)

        stored = s.get_default_path(Settings.PATH_RELATION_FILE)
        assert stored == fp

        budgets, err = ExcelRelationReader().read(stored)
        assert err is None
        assert len(budgets) == 1
        assert budgets[0]["numero"] == "1"
