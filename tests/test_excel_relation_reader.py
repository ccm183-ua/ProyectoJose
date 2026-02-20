"""
Tests para ExcelRelationReader.

Cubren:
- Lectura de archivo válido
- Conversión de fechas datetime → DD-MM-YY
- Salto de filas de cabecera y filas vacías
- Manejo de archivo inexistente y formato inválido
- Compatibilidad de salida con ProjectParser
"""

import os
import tempfile
import shutil
from datetime import datetime

import pytest
from openpyxl import Workbook

from src.core.excel_relation_reader import ExcelRelationReader


@pytest.fixture
def temp_dir():
    path = tempfile.mkdtemp()
    yield path
    shutil.rmtree(path, ignore_errors=True)


def _create_relation_excel(path, rows, header_row=3, company_name="Test S.L."):
    """Helper: crea un xlsx con la estructura de relación de presupuestos.

    ``rows`` es una lista de tuplas con valores para las columnas A–J.
    Los datos se escriben a partir de ``header_row + 1``.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=company_name)
    ws.cell(row=2, column=1, value="RELACIÓN DE PRESUPUESTOS EMITIDOS 2026")
    headers = ["Nº", "FECHA", "CLIENTE", "MEDIACIÓN", "CALLE",
               "NUM", "C.P", "LOCALIDAD", "TIPO", "IMPORTE"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    for r_idx, row_data in enumerate(rows, start=header_row + 1):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


class TestReadValidFile:
    """Lectura de archivos válidos."""

    def test_read_valid_file(self, temp_dir):
        """Lee un archivo con 2 filas de datos y devuelve 2 dicts."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "Cliente A", "MED", "C/ Mayor", "12", "03001", "Alicante", "Reforma", 1500),
            (2, datetime(2026, 2, 15), "Cliente B", "", "C/ Sol", "5", "", "Madrid", "Pintura", 800),
        ])
        reader = ExcelRelationReader()
        budgets, err = reader.read(fp)
        assert err is None
        assert len(budgets) == 2

    def test_read_returns_correct_keys(self, temp_dir):
        """Cada dict tiene las claves esperadas."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "CLI", "MED", "C/", "1", "03001", "LOC", "TIPO", 100),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        expected_keys = {"numero", "fecha", "cliente", "mediacion", "calle",
                         "num_calle", "codigo_postal", "localidad", "tipo", "importe"}
        assert set(budgets[0].keys()) == expected_keys

    def test_read_all_rows_parsed(self, temp_dir):
        """Todas las filas con número se parsean."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        rows = [(i, datetime(2026, 1, i), f"C{i}", "", "C/", str(i), "", "L", "T", 0)
                for i in range(1, 11)]
        _create_relation_excel(fp, rows)
        budgets, _ = ExcelRelationReader().read(fp)
        assert len(budgets) == 10

    def test_read_importe_included(self, temp_dir):
        """El campo importe se incluye correctamente."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "CLI", "", "C/", "1", "", "", "T", 4650),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert budgets[0]["importe"] == "4650.0"


class TestDateConversion:
    """Conversión de fechas."""

    def test_read_date_conversion(self, temp_dir):
        """datetime se convierte a DD-MM-YY."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "C", "", "C/", "1", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert budgets[0]["fecha"] == "08-01-26"

    def test_read_handles_string_date(self, temp_dir):
        """Si la fecha ya es string, se deja tal cual."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, "15-03-26", "C", "", "C/", "1", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert budgets[0]["fecha"] == "15-03-26"

    def test_read_handles_mixed_date_formats(self, temp_dir):
        """Mezcla de datetime y strings se manejan ambas."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 3, 1), "C1", "", "C/", "1", "", "", "", None),
            (2, "10-04-26", "C2", "", "C/", "2", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert budgets[0]["fecha"] == "01-03-26"
        assert budgets[1]["fecha"] == "10-04-26"


class TestRowFiltering:
    """Filtrado de filas no válidas."""

    def test_read_skips_header_rows(self, temp_dir):
        """Las filas de cabecera y título no se incluyen."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 1), "C", "", "C/", "1", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert len(budgets) == 1
        assert budgets[0]["numero"] == "1"

    def test_read_skips_empty_rows(self, temp_dir):
        """Filas completamente vacías se ignoran."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 1), "C", "", "C/", "1", "", "", "", None),
            (None, None, None, None, None, None, None, None, None, None),
            (3, datetime(2026, 1, 3), "C3", "", "C/", "3", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert len(budgets) == 2

    def test_read_skips_rows_without_number(self, temp_dir):
        """Filas sin número en columna A se ignoran."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 1), "C", "", "C/", "1", "", "", "", None),
            ("", datetime(2026, 1, 2), "SinNum", "", "C/", "2", "", "", "", None),
            (3, datetime(2026, 1, 3), "C3", "", "C/", "3", "", "", "", None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        assert len(budgets) == 2


class TestMissingOptionalFields:
    """Manejo de campos opcionales vacíos."""

    def test_read_handles_missing_optional_fields(self, temp_dir):
        """Campos opcionales vacíos se devuelven como string vacío."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "CLI", None, "C/", None, None, None, None, None),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        b = budgets[0]
        assert b["mediacion"] == ""
        assert b["num_calle"] == ""
        assert b["codigo_postal"] == ""
        assert b["localidad"] == ""
        assert b["tipo"] == ""
        assert b["importe"] == ""


class TestErrorHandling:
    """Manejo de errores."""

    def test_read_file_not_found(self):
        """Archivo inexistente devuelve error."""
        budgets, err = ExcelRelationReader().read("/no/existe/archivo.xlsx")
        assert budgets == []
        assert err is not None
        assert "no encontrado" in err.lower() or "not found" in err.lower() or "no encontrado" in err.lower()

    def test_read_invalid_format(self, temp_dir):
        """Archivo sin cabeceras válidas devuelve error."""
        fp = os.path.join(temp_dir, "bad.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Datos aleatorios")
        ws.cell(row=2, column=1, value="Sin cabeceras Nº")
        wb.save(fp)
        wb.close()
        budgets, err = ExcelRelationReader().read(fp)
        assert budgets == []
        assert "cabecera" in err.lower() or "Nº" in err


class TestCompatibility:
    """Compatibilidad con ProjectParser."""

    def test_output_compatible_with_project_parser(self, temp_dir):
        """Las claves de salida coinciden con las de ProjectParser.parse_clipboard_data."""
        fp = os.path.join(temp_dir, "rel.xlsx")
        _create_relation_excel(fp, [
            (1, datetime(2026, 1, 8), "CLI", "MED", "C/ Test", "10", "03001", "Alicante", "Reforma", 500),
        ])
        budgets, _ = ExcelRelationReader().read(fp)
        b = budgets[0]

        from src.core.project_parser import ProjectParser
        parser = ProjectParser()
        tsv = "1\t08-01-26\tCLI\tMED\tC/ Test\t10\t03001\tAlicante\tReforma"
        parsed, parse_err = parser.parse_clipboard_data(tsv)
        assert parse_err is None

        for key in parsed:
            assert key in b, f"Clave '{key}' de ProjectParser falta en ExcelRelationReader"
